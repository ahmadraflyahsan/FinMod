"""Export utilities: Excel + PDF"""
import io, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter
from datetime import date

def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def _hdr(ws, row, col, val, bg="1F4E79", fg="FFFFFF", bold=True, wrap=True):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = _fill(bg); c.border = _thin()
    c.font = Font(bold=bold, color=fg, name="Calibri", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    return c

def _num(ws, row, col, val, fmt='#,##0', bg=None):
    c = ws.cell(row=row, column=col, value=round(float(val),2) if val else 0)
    c.number_format = fmt
    c.border = _thin()
    c.font = Font(name="Calibri", size=9)
    c.alignment = Alignment(horizontal="right", vertical="center")
    if bg: c.fill = _fill(bg)
    if isinstance(val,(int,float)) and val<0:
        c.font = Font(name="Calibri", size=9, color="C0392B")
    return c

def _label(ws, row, col, val, indent=0, bold=False, bg=None):
    c = ws.cell(row=row, column=col, value=("  "*indent)+str(val))
    c.border = _thin()
    c.font = Font(name="Calibri", size=9, bold=bold)
    c.alignment = Alignment(horizontal="left", vertical="center")
    if bg: c.fill = _fill(bg)
    return c

def build_cashflow_excel(project, all_results):
    """Build comprehensive Excel with all sheets"""
    wb = Workbook()
    units = project.get("units",[])
    years = all_results[0]["years"] if all_results else []

    # ── SUMMARY sheet ──────────────────────────────────────────────────────
    ws = wb.active; ws.title = "Summary"
    ws.merge_cells("A1:F1")
    c = ws.cell(1,1,f"GEOTHERMAL FINANCIAL MODEL – {project.get('project_name','')}")
    c.font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    c.fill = _fill("0D3B6E"); c.alignment = Alignment(horizontal="center")

    headers = ["Unit","Scheme","Capacity (MW)","COD","NPV UFCF (M USD)","IRR UFCF (%)","NPV LFCF (M USD)","IRR LFCF (%)","Payback (yr)","LCOE (¢/kWh)","Verdict"]
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 3, ci, h, bg="1A5276")
    for i, res in enumerate(all_results):
        u = units[i] if i < len(units) else {}
        row = 4+i
        npv = res.get("npv_ufcf",0); irr = res.get("irr_ufcf"); hr = float(u.get("hurdle_rate") or 0.076)
        verdict = "✅ Viable" if npv>0 and irr and irr>hr else "⚠️ Marginal" if npv>-50000 else "❌ Not Viable"
        vals = [u.get("unit_name",f"Unit {i+1}"), u.get("scheme","PJBL"),
                u.get("install_capacity",""), str(u.get("cod_date","")),
                round(npv/1000,1), round(irr*100,2) if irr else None,
                round(res.get("npv_lfcf",0)/1000,1),
                round(res.get("irr_lfcf",0)*100,2) if res.get("irr_lfcf") else None,
                res.get("payback"), round(res.get("lcoe",0),2), verdict]
        for ci, v in enumerate(vals,1):
            c = ws.cell(row,ci,v); c.border = _thin()
            c.font = Font(name="Calibri",size=9)
            c.alignment = Alignment(horizontal="center",vertical="center")
    ws.column_dimensions["A"].width = 15
    for ci in range(2,12): ws.column_dimensions[get_column_letter(ci)].width = 14

    # ── Annual Model per Unit ───────────────────────────────────────────────
    for i, res in enumerate(all_results):
        u = units[i] if i < len(units) else {}
        uname = u.get("unit_name", f"Unit {i+1}")
        ws2 = wb.create_sheet(f"{uname[:20]}_CF")

        # Title
        ws2.merge_cells(f"A1:{get_column_letter(len(years)+2)}1")
        c = ws2.cell(1,1,f"Annual Cash Flow – {uname} | {project.get('project_name','')}")
        c.font = Font(bold=True,size=11,color="FFFFFF",name="Calibri")
        c.fill = _fill("0D3B6E"); c.alignment = Alignment(horizontal="center")

        # Year headers
        _hdr(ws2,2,1,"Parameter",bg="2C3E50"); _hdr(ws2,2,2,"Unit",bg="2C3E50")
        for ci, yr in enumerate(years,3):
            _hdr(ws2,2,ci,str(yr),bg="1A5276")

        # Sections
        SECTIONS = [
            # (label, key, unit_str, indent, bold, bg)
            ("WELL COUNT","","Well",0,True,"2C3E50"),
            ("Net Generation","net_gen","MWh",0,False,None),
            ("Gross Revenue","revenue","kUSD",0,True,"1E8449"),
            ("","","",0,False,None),
            ("CAPEX","capex","kUSD",0,True,"922B21"),
            ("  Exploration Phase","_capex_exp","kUSD",1,False,None),
            ("  Development Phase","_capex_dev","kUSD",1,False,None),
            ("  Operation (Makeup)","_capex_ops","kUSD",1,False,None),
            ("","","",0,False,None),
            ("OPEX","opex","kUSD",0,True,"A04000"),
            ("  O&M Steam Field","_om_steam","kUSD",1,False,None),
            ("  O&M Power Plant","_om_plant","kUSD",1,False,None),
            ("  Overhaul","_overhaul","kUSD",1,False,None),
            ("  Chemical Treatment","_chem","kUSD",1,False,None),
            ("  Transmisi","_trans","kUSD",1,False,None),
            ("","","",0,False,None),
            ("EBITDA","ebitda","kUSD",0,True,"1A5276"),
            ("Depreciation","depreciation","kUSD",0,False,None),
            ("EBIT","ebit","kUSD",0,True,"1A5276"),
            ("Production Fee","prod_fee","kUSD",0,False,None),
            ("Production Bonus","prod_bonus","kUSD",0,False,None),
            ("Income Tax","income_tax","kUSD",0,False,None),
            ("","","",0,False,None),
            ("UFCF (Unlevered FCF)","ufcf","kUSD",0,True,"1E6B45"),
            ("  Cum. UFCF","cum_ufcf","kUSD",1,False,None),
            ("  Disc. UFCF","disc_ufcf","kUSD",1,False,None),
            ("  Cum. Disc. UFCF","cum_disc_ufcf","kUSD",1,False,None),
            ("","","",0,False,None),
            ("Interest Expense","interest","kUSD",0,False,None),
            ("Debt Repayment","debt_repay","kUSD",0,False,None),
            ("LFCF (Levered FCF)","lfcf","kUSD",0,True,"6C3483"),
            ("  Cum. LFCF","cum_lfcf","kUSD",1,False,None),
            ("  Disc. LFCF","disc_lfcf","kUSD",1,False,None),
            ("  Cum. Disc. LFCF","cum_disc_lfcf","kUSD",1,False,None),
            ("","","",0,False,None),
            ("Discount Factor","disc_factor","x",0,False,None),
        ]

        row = 3
        for sec in SECTIONS:
            label, key, unit_str, indent, bold, bg = sec
            bg_hex = bg if bg else ("F7F9FC" if row%2==0 else "FFFFFF")
            _label(ws2, row, 1, label, indent, bold, bg_hex)
            _label(ws2, row, 2, unit_str, 0, False, bg_hex)
            if key and key in res:
                data = res[key]
                for ci, yr in enumerate(years, 3):
                    v = data.get(yr, 0) if isinstance(data,dict) else 0
                    fmt = '#,##0' if 'factor' not in key else '0.000'
                    _num(ws2, row, ci, v, fmt, bg_hex)
            else:
                for ci, _ in enumerate(years, 3):
                    c = ws2.cell(row,ci,""); c.border=_thin()
                    if bg: c.fill = _fill(bg)
            row += 1

        # Widths
        ws2.column_dimensions["A"].width = 32
        ws2.column_dimensions["B"].width = 8
        ws2.row_dimensions[1].height = 22
        ws2.row_dimensions[2].height = 30
        for ci in range(3, len(years)+3):
            ws2.column_dimensions[get_column_letter(ci)].width = 11
        ws2.freeze_panes = "C3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def build_summary_pdf(project, all_results):
    """Simple PDF summary using reportlab"""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer, HRFlowable)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
    except ImportError:
        return None

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    H1 = ParagraphStyle("H1", parent=styles["Heading1"],
                        fontSize=16, textColor=colors.HexColor("#0D3B6E"),
                        spaceAfter=6)
    H2 = ParagraphStyle("H2", parent=styles["Heading2"],
                        fontSize=11, textColor=colors.HexColor("#1A5276"),
                        spaceAfter=4)
    BODY = ParagraphStyle("Body", parent=styles["Normal"],
                          fontSize=8, spaceAfter=2)

    story = []
    story.append(Paragraph(f"Geothermal Financial Model – {project.get('project_name','')}", H1))
    story.append(Paragraph(f"PT Pertamina Geothermal Energy | Generated: {date.today():%d %B %Y}", BODY))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#0D3B6E")))
    story.append(Spacer(1, 0.3*cm))

    # Summary table
    units = project.get("units",[])
    data = [["Unit","Scheme","Cap (MW)","COD","NPV UFCF\n(M USD)","IRR UFCF\n(%)","NPV LFCF\n(M USD)","Payback\n(yr)","LCOE\n(¢/kWh)","Verdict"]]
    for i, res in enumerate(all_results):
        u = units[i] if i<len(units) else {}
        npv = res.get("npv_ufcf",0); irr = res.get("irr_ufcf"); hr = float(u.get("hurdle_rate") or 0.076)
        verdict = "Viable" if npv>0 and irr and irr>hr else "Marginal"
        data.append([u.get("unit_name",f"Unit {i+1}"), u.get("scheme","—"),
                     f"{u.get('install_capacity','—')} MW", str(u.get("cod_date","—")),
                     f"{npv/1000:,.1f}", f"{irr*100:.2f}%" if irr else "N/A",
                     f"{res.get('npv_lfcf',0)/1000:,.1f}",
                     str(res.get("payback","—")), f"{res.get('lcoe',0):.2f}", verdict])

    col_w = [3*cm,2*cm,2*cm,2.5*cm,2.5*cm,2*cm,2.5*cm,2*cm,2*cm,2*cm]
    t = Table(data, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#0D3B6E")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,-1),"Helvetica"),
        ("FONTSIZE",(0,0),(-1,0),8),("FONTSIZE",(0,1),(-1,-1),7),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#EBF5FB")]),
        ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#BDC3C7")),
        ("BOTTOMPADDING",(0,0),(-1,0),6),
    ]))
    story.append(t)
    story.append(Spacer(1,0.4*cm))

    # Per-unit KPI cards
    story.append(Paragraph("Economic Indicators – Per Unit", H2))
    for i, res in enumerate(all_results):
        u = units[i] if i<len(units) else {}
        irr_u = res.get("irr_ufcf"); irr_l = res.get("irr_lfcf")
        kpi_data = [
            ["Indicator","UFCF (Unlevered)","LFCF (Levered)"],
            ["NPV (M USD)", f"{res.get('npv_ufcf',0)/1000:,.2f}", f"{res.get('npv_lfcf',0)/1000:,.2f}"],
            ["IRR (%)", f"{irr_u*100:.2f}%" if irr_u else "N/A", f"{irr_l*100:.2f}%" if irr_l else "N/A"],
            ["Payback (yr)", str(res.get("payback","—")), str(res.get("payback_lfcf","—"))],
            ["LCOE (¢/kWh)", f"{res.get('lcoe',0):.3f}", "—"],
            ["Total CAPEX (M USD)", f"{res.get('total_capex',0)/1000:,.1f}", "—"],
            ["Total OPEX (M USD)", f"{res.get('total_opex',0)/1000:,.1f}", "—"],
            ["Net Generation (GWh)", f"{res.get('total_gen_gwh',0):,.1f}", "—"],
        ]
        story.append(Paragraph(f"▸ {u.get('unit_name',f'Unit {i+1}')} | {u.get('scheme','—')} | "
                               f"{u.get('install_capacity','—')} MW | COD: {u.get('cod_date','—')}", BODY))
        kt = Table(kpi_data, colWidths=[5*cm,4*cm,4*cm])
        kt.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1A5276")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTNAME",(0,0),(-1,-1),"Helvetica"),
            ("FONTSIZE",(0,0),(-1,-1),7),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F0F4F8")]),
            ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#BDC3C7")),
            ("ALIGN",(1,0),(-1,-1),"CENTER"),
        ]))
        story.append(kt)
        story.append(Spacer(1,0.3*cm))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()
